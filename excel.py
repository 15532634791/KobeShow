# -*- coding:utf8 -*-
from openpyxl import load_workbook
import os
from ruamel import yaml
import re


class Excel(object):
    def __init__(self,file):
        self.file = os.path.join(os.getcwd(), file)
        self.workbook = load_workbook(self.file)
        self.sheet_names = self.workbook.sheetnames

    def work_sheet(self,work_sheet=None):
        self.sheet = self.workbook[work_sheet]
        return self.sheet

    def cell(self,row,col):
        try:
            sheet = self.sheet
        except AttributeError:
            raise NotImplementedError('object.work_sheet(sheet_name)')
        if sheet.cell(row=row,column=col).value:
            return sheet.cell(row=row,column=col).value
        else:
            return None


def transfer_list(arg):
    if len(str(arg).split(',')) > 1:
        arg_value = []
        for num in str(arg).split(','):
            value = re.findall('(\d+)', num)[0]
            arg_value.append(int(value))
        return arg_value
    else:
        return arg


def get_range(sheet):
    handle_range = list()
    row_value = 0
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row,column=1).value:
            handle_range.append((row_value,row - 1))
            row_value = row
    handle_range.append((row_value,sheet.max_row))
    return handle_range[1:]


#object参数为实例化的对象名称，start,end为区间范围
def title(object,start,end):
    title = {}
    for row in range(start,end):
        if object.cell(row,1):
            title.update({'country':object.cell(row,1)})
        if object.cell(row,2):
            title.update({'area': object.cell(row, 2)})
        if object.cell(row,3):
            title.update({'flags': object.cell(row, 3)})
        if object.cell(row,4):
            title.update({'phone': object.cell(row, 4)})
        if object.cell(row,5):
            title.update({'model': object.cell(row, 5)})
        if object.cell(row,6):
            title.update({'stack': object.cell(row, 6)})
    return title


#object参数为实例化的对象名称，start,end为区间范围
def famous(object,start,end):
    port_tag = {}
    port_type_cache = ''
    for row in range(start,end):

        port_type_value = {'balls': transfer_list(object.cell(row, 8)),'level': object.cell(row, 9),
                            'speed':transfer_list(object.cell(row, 10)),

                           }

        port_type = object.cell(row, 7)
        if port_type:
            port_tag.update({port_type : port_type_value})
            port_type_cache = port_type
        else:
            for key,value in port_type_value.items():
                if value:
                    previous_value = port_tag.get(port_type_cache).get(key)
                    if isinstance(previous_value,str) or isinstance(previous_value,int):
                        temp = [previous_value]
                        temp.append(value)
                        port_tag[port_type_cache][key] = temp

                    if isinstance(previous_value,list):
                        temp = previous_value
                        temp.append(value)
                        port_tag[port_type_cache][key] = temp

    port_tag ={'famous':port_tag}
    return port_tag


#主函数负责处理数据
def main():
    # 打开excel
    excel = Excel('球员信息.xlsx')
    # 获取sheet页名称
    # sheet_name = excel.sheet_names[0]
    #遍历sheet页，获取所有sheet页信息
    for sheet_name in excel.sheet_names:
        # 获取sheet页信息
        sheet = excel.work_sheet(sheet_name)
        title_list = list()
        for interval in get_range(sheet):
            # print(interval)
            #获取端口区间
            start,end =interval
            #获取title信息
            get_title = title(excel,start,end)
            #获取port_tag信息
            get_port_tag = famous(excel,start,end)
            #合并title、port_tag
            get_title.update(get_port_tag)
            # print(get_title)
            title_list.append(get_title)
        # print(yaml.dump(title_list, Dumper=yaml.RoundTripDumper))

        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),'file')
        file_name = '{}.yaml'.format(sheet_name)
        name = os.path.join(file_path,file_name)

        file = open(name, 'w', encoding='utf8')
        file.write(yaml.dump(title_list, Dumper=yaml.RoundTripDumper))
        file.close()



if __name__ == '__main__':
    main()

