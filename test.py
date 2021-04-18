# -*- coding:utf8 -*-
import os
import sys
from openpyxl import load_workbook


class Cloud(object):

    def __init__(self,excel_name):
        try:
            #python2使用
            reload(sys)
            # 设置默认编码方式为：UTF8，其他编码方式会报错
            sys.setdefaultencoding('utf-8')

        except NameError:
            pass

        self.excel_name = excel_name
        # 打开excel(工作薄)
        self.workbook = load_workbook(self.excel_name)

    def sheet(self):
        sheet = self.workbook['最新题目']
        return sheet

    @classmethod
    def create_file(cls,value):
        with open('AlibabaCloudComputing.txt','a+',encoding='utf8') as f:
            f.write(value)
            f.write('\n')

    def run(self):
        sheet = self.sheet()

        if os.path.isfile('AlibabaCloudComputing.txt'):
            os.remove('AlibabaCloudComputing.txt')
            self.create_file('阿里云计算题库\n')
        else:
            self.create_file('阿里云计算题库\n')

        for row in range(2,sheet.max_row+1):

            topic_type = sheet.cell(row,2).value.strip()
            topic_name = '{} {}'.format(row-1,sheet.cell(row,3).value.strip())
            topic_answer = '{} {}'.format('答案',sheet.cell(row,4).value.strip())
            topic_analysis = '{} {}'.format('题目解析',sheet.cell(row,6).value)
            option_a = "{} {}".format('A',sheet.cell(row,8).value)
            option_b = "{} {}".format('B',sheet.cell(row,9).value)
            option_c = "{} {}".format('C',sheet.cell(row,10).value)
            option_d = "{} {}".format('D',sheet.cell(row,11).value)
            option_e = "{} {}".format('E',sheet.cell(row,12).value)
            option_f = "{} {}".format('F',sheet.cell(row,13).value)
            option_g = "{} {}".format('G',sheet.cell(row,14).value)

            self.create_file(topic_name)
            self.create_file(topic_type)
            if sheet.cell(row,8).value:
                self.create_file(option_a)
            if sheet.cell(row,9).value:
                self.create_file(option_b)
            if sheet.cell(row,10).value:
                self.create_file(option_c)
            if sheet.cell(row,11).value:
                self.create_file(option_d)
            if sheet.cell(row,12).value:
                self.create_file(option_e)
            if sheet.cell(row,13).value:
                self.create_file(option_f)
            if sheet.cell(row,14).value:
                self.create_file(option_g)
            self.create_file(topic_answer)
            self.create_file(topic_analysis)
            self.create_file('\r')


if __name__ == '__main__':
    cloud = Cloud('AlibabaCloudComputing.xlsx')
    cloud.run()
